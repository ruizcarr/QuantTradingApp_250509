#IMPORT RELEVANT MODULES

#Import libraries and functions
import numpy as np,random


np.random.seed(42)
random.seed(42)
import pandas as pd
import matplotlib.pyplot as plt
import datetime
import time
import os.path
import pickle
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score

# Wider print limits
pd.set_option("display.max_columns", None)
pd.set_option('display.width', None)
# Silence warnings
import warnings
warnings.filterwarnings('ignore')


#from Backtest_Vectorized import compute_backtest_vectorized
from Markowitz_Vectorized import compute_optimized_markowitz_d_w
from WalkForwardTraining import WalkForwardTraining
import Market_Data_Feed as mdf
from utils import mean_positions


#Get SETTINGS
from config import settings,utils
settings=settings.get_settings() #Edit Settings Dict at file config/settings.py

# MAIN CODE
def run(settings):

    #DATA & INDICATORS
    data_ind=mdf.Data_Ind_Feed(settings).data_ind
    data, indicators_dict = data_ind
    tickers_returns=data.tickers_returns

    #Compute DDN Training and save results to Jason
    ddn_window = 17
    nd = 22  # 22 n days forward returns calculation
    compute_ddn_weight_training(tickers_returns,ddn_window, nd, study=True)


    #Compute DDN Weights Time Series
    ddn_weight_ts= compute_ddn_weight_ts(tickers_returns, study=True)


    #Back Test

    #Retrive System Positions
    folder_name = 'results'
    csv_filename2 = 'training_positions.csv'
    full_path2 = os.path.join(folder_name, csv_filename2)
    positions = pd.read_csv(full_path2, index_col=0)
    positions.index = pd.to_datetime(positions.index)
    
    #Original System Returns
    ret=tickers_returns.reindex(positions.index)
    pos_ret=positions*ret

    ddn_weight_ts = ddn_weight_ts.reindex(positions.index)
    ddn_ret=ddn_weight_ts*pos_ret
    
    #Add sum
    pos_ret['sum']=pos_ret.sum(axis=1)
    ddn_ret['sum'] = ddn_ret.sum(axis=1)

    #Cumulated Returns
    pos_cumret = (1 + pos_ret).cumprod()
    ddn_cumret=(1+ddn_ret).cumprod()


    ddn_cumret.plot()

    plot_df2= pd.DataFrame()
    plot_df2['pos_cumret_sum']=pos_cumret['sum']
    plot_df2['ddn_cumret_sum'] = ddn_cumret['sum']
    plot_df2.plot()

    ddn_sum,_=compute_ddn(plot_df2.pct_change(),ddn_window=250)

    ddn_sum.plot()



    plt.show()

def compute_ddn_weight_training(
                                tickers_returns,
                                ddn_window=22*3,
                                nd=10, #22 n days forward returns calculation
                                study=True):

    percentiles = [0.025, 0.05, 0.1, 0.2, 0.3, 0.40, 0.50, 0.60, 0.70, 0.80, 0.90,1]

    ddn,cum_ret =compute_ddn(tickers_returns.shift(1),ddn_window=ddn_window)

    #Read Positions
    positions=get_training_positions_from_csv()

    ddn=ddn.reindex(positions.index)

    # Run analysis
    results = optimal_buy_strategy_ndays(tickers_returns, ddn, positions, percentiles, nd=nd)
    ddn_levels, kelly_pcts=create_df_report(results)

    #Create Drawdown Weight
    kelly_clip = kelly_pcts.clip(upper=1.10 * kelly_pcts.iloc[-1], axis=1) / 100  # limited to n pct over value at trend (ddn=0) 0-0.5
    kelly_clip = kelly_clip.clip(upper=0.40) # 0-0.35
    kelly_norm=kelly_clip/kelly_clip.mean().max() #normalized by max of means, values 0-1.2
    ddn_weight = kelly_norm#-0.05#/kelly_norm.mean().mean() #values 0-2
    ddn_weight = ddn_weight.clip(upper=1.25, lower=0.5)#values 0-1.25

    ddn_study_results = {
        'ddn_levels': ddn_levels,
        'kelly_pcts': kelly_pcts,
        'ddn_weight': ddn_weight,
        'percentiles': percentiles,
        'ddn_window': ddn_window,
    }

    #Save to Jason
    os.makedirs('trained_models', exist_ok=True)
    with open('trained_models/ddn_study_results.pkl', 'wb') as f:
        pickle.dump(ddn_study_results, f)

    #To Load
    #with open('results/ddn_study_results.pkl', 'rb') as f:
    #    my_dict = pickle.load(f)
    # RETRIEVE individual items from the dict
    #ddn_levels = my_dict['ddn_levels']
    #kelly_pcts = my_dict['kelly_pcts']
    #ddn_weight = my_dict['ddn_weight']
    #percentiles = my_dict['percentiles']

    if study:
        ddn_pos=ddn.where(positions>0)
        ddn_pos.fillna(method='ffill', inplace=True)

        for ticker in tickers_returns.columns:
            plot_df = pd.DataFrame()
            plot_df['positions*10'] = positions[ticker]*10
            plot_df['cum_ret'] = cum_ret[ticker]
            plot_df['ddn%'] = ddn[ticker]*100
            plot_df['ddn_pos%'] = ddn_pos[ticker] * 100
            plot_df.plot(title=ticker)

        # Create Excel report
        create_excel_report(results, 'results/optimal_buy_strategy_ndays.xlsx')

    return

def compute_ddn_weight_ts(tickers_returns, study=True):
    # RETRIEVE
    with open('trained_models/ddn_study_results.pkl', 'rb') as f:
        my_dict = pickle.load(f)
    # RETRIEVE individual items from the dict
    ddn_levels = my_dict['ddn_levels']
    kelly_pcts = my_dict['kelly_pcts']
    ddn_weight = my_dict['ddn_weight']
    ddn_window = my_dict['ddn_window']

    # Compute DDN Index Time Series
    ddn, _ = compute_ddn(tickers_returns.shift(1), ddn_window)
    ddn_weight_ts = map_ddn_to_weights_ts(ddn, ddn_levels, ddn_weight)

    if study:
        x = ddn_levels
        x_name = 'ddn_levels'
        y = kelly_pcts
        y_name = 'kelly_pcts'
        plot_x_y(x, x_name, y, y_name)

        y = ddn_weight
        y_name = 'ddn_weight'
        plot_x_y(x, x_name, y, y_name)

        ddn_weight_ts.plot(title='ddn_weight_ts')

    return ddn_weight_ts

def map_ddn_to_weights_ts(ddn: pd.DataFrame,
                     ddn_levels: pd.DataFrame,
                     ddn_weight: pd.DataFrame) -> pd.DataFrame:
    """
    Map ddn timeseries to ddn_weight values

    For each point in ddn timeseries, find which ddn_level it falls into,
    and return the corresponding ddn_weight value.

    Args:
        ddn: DataFrame with ddn timeseries (rows=time, cols=assets)
             Shape: (n_periods, n_assets)
        ddn_levels: DataFrame with ddn thresholds by percentile (rows=percentile_idx, cols=assets)
                    Shape: (11, n_assets)
        ddn_weight: DataFrame with expected returns by regime (rows=percentile_idx, cols=assets)
                   Shape: (11, n_assets) - same index as ddn_levels

    Returns:
        ddn_weight_time_series: DataFrame with ddn_weight values for each time point
                               Same shape as ddn (n_periods, n_assets)
                               Each value is the ddn_weight corresponding to that ddn value

    Example:
        ddn has shape (250, 6) - 250 days, 6 assets
        ddn_levels has shape (11, 6) - 11 percentile levels, 6 assets
        ddn_weight has shape (11, 6) - expected returns for each level

        Output: (250, 6) - for each day and asset, the ddn_weight value
                           corresponding to that day's ddn level
    """
    ddn_weight_ts = pd.DataFrame(
        index=ddn.index,
        columns=ddn.columns,
        dtype=float
    )

    for asset in ddn.columns:
        # Get the levels and index for this asset, sorted ascending
        levels = ddn_levels[asset].sort_values()
        levels_idx = ddn_levels[asset].index[ddn_levels[asset].argsort()]

        # For each time point in ddn
        for t in ddn.index:
            current_ddn = ddn.loc[t, asset]

            # Find which level this ddn falls into
            # Count how many levels are <= current ddn
            level_count = (levels <= current_ddn).sum()

            # Get the corresponding index row
            if level_count == 0:
                # Below all levels (shouldn't happen if ddn_levels covers range)
                regime_idx = levels_idx[0]
            elif level_count >= len(levels):
                # At or above all levels
                regime_idx = levels_idx[-1]
            else:
                # Use the level we just crossed
                regime_idx = levels_idx[level_count - 1]

            # Get the ddn_weight value for this regime
            ddn_weight_ts.loc[t, asset] = ddn_weight.loc[regime_idx, asset]

    ddn_weight_ts.index = pd.to_datetime(ddn_weight_ts.index)

    #ddn_weight_ts=ddn_weight_ts.rolling(3).mean()

    return ddn_weight_ts


def compute_ddn(tickers_returns,ddn_window=250):
    cum_ret=(1+tickers_returns).cumprod()
    rolling_max = cum_ret.rolling(window=ddn_window, min_periods=1).max()
    ddn= (cum_ret / rolling_max) - 1
    return ddn,cum_ret

def get_training_positions_from_csv():
    import os
    folder_name = 'results'
    csv_filename2 = 'training_positions.csv'
    full_path2 = os.path.join(folder_name, csv_filename2)
    positions = pd.read_csv(full_path2, index_col=0)
    positions.index = pd.to_datetime(positions.index)
    return positions

def plot_x_y(x_df,x_name,y_df,y_name):
    # Create subplots for each column
    columns = x_df.columns
    n_cols = 3
    n_rows = (len(columns) + n_cols - 1) // n_cols

    fig, axes = plt.subplots(n_rows, n_cols, figsize=(15, 10))
    axes = axes.flatten()

    for idx, col in enumerate(columns):
        ax = axes[idx]

        x = x_df[col]
        y = y_df[col]
        valid_mask = ~(x.isna() | y.isna())

        x_clean = x[valid_mask]
        y_clean = y[valid_mask]

        if len(x_clean) > 0:
            # Sort by x values (so zero is at the left)
            sorted_indices = np.argsort(x_clean.values)[::-1]
            x_sorted = x_clean.iloc[sorted_indices]
            y_sorted = y_clean.iloc[sorted_indices]

            ax.plot(x_sorted, y_sorted, marker='o', linestyle='-', linewidth=2, markersize=6, label=col)
            ax.set_xlabel('DDN Levels', fontsize=10)
            ax.set_ylabel('Kelly %', fontsize=10)
            ax.set_title(f'{col}', fontsize=11, fontweight='bold')
            ax.grid(True, alpha=0.3)
            ax.legend()
        else:
            ax.text(0.5, 0.5, 'No Data', ha='center', va='center', transform=ax.transAxes)
            ax.set_title(f'{col}', fontsize=11, fontweight='bold')

    # Remove extra subplots
    for idx in range(len(columns), len(axes)):
        fig.delaxes(axes[idx])




def optimal_buy_strategy_ndays(tickers_returns, ddn_df, positions_df,percentiles,nd=5):
    """
    Analyze n-day returns with Kelly criterion and strategy returns
    """

    asset_columns = tickers_returns.columns.tolist()

    all_results = {}

    for asset in asset_columns:
        if asset not in ddn_df.columns or asset not in positions_df.columns:
            continue

        asset_ddn = ddn_df[asset].dropna()
        returns_col = tickers_returns[asset]
        positions_col = positions_df[asset]

        thresholds = {
            f'{int(p * 100)}th_percentile': asset_ddn.quantile(p)
            for p in percentiles
        }

        asset_performance = {}

        for threshold_name, ddn_level in thresholds.items():
            buy_conditions = (ddn_df[asset] <= ddn_level) & (positions_df[asset] > 0.01) #& (ddn_df[asset]>ddn_df[asset].rolling(5).mean())
            buy_signals = ddn_df[buy_conditions].index

            if len(buy_signals) == 0:
                asset_performance[threshold_name] = {
                    'ddn_level': ddn_level,
                    'count': 0,
                    'avg_pos_size': np.nan,
                    'avg_return_nd': np.nan,
                    'avg_win_nd': np.nan,
                    'avg_loss_nd': np.nan,
                    'avg_max_loss_nd': np.nan,
                    'win_rate_nd': np.nan,
                    'risk_reward_ratio': np.nan,
                    'kelly_percentage': np.nan,
                    'strategy_return': np.nan,
                    'annualized_return': np.nan
                }
                continue

            returns_nd = []
            position_sizes = []

            for buy_date in buy_signals:
                try:
                    idx_pos = tickers_returns.index.get_loc(buy_date)
                except KeyError:
                    continue

                position_sizes.append(positions_df.loc[buy_date, asset])

                future_nd = tickers_returns.iloc[idx_pos + 1:idx_pos + nd+1][asset]
                if len(future_nd) > 0:
                    ret_nd = (future_nd.sum()) * 100
                    returns_nd.append(ret_nd)

            if len(returns_nd) == 0:
                asset_performance[threshold_name] = {
                    'ddn_level': ddn_level,
                    'count': len(buy_signals),
                    'avg_pos_size': np.mean(position_sizes) if position_sizes else np.nan,
                    'avg_return_nd': np.nan,
                    'avg_win_nd': np.nan,
                    'avg_loss_nd': np.nan,
                    'avg_max_loss_nd': np.nan,
                    'win_rate_nd': np.nan,
                    'risk_reward_ratio': np.nan,
                    'kelly_percentage': np.nan,
                    'strategy_return': np.nan,
                    'annualized_return': np.nan
                }
                continue

            returns_array = np.array(returns_nd)
            wins = returns_array[returns_array > 0]
            losses = returns_array[returns_array < 0]

            win_count = len(wins)
            loss_count = len(losses)
            win_rate = (win_count / len(returns_array) * 100) if len(returns_array) > 0 else 0

            #Shrinkage: Ammended winrate by number of observations
            n=win_count+loss_count
            win_rate = max(0,(win_rate/100-1/n**0.5)*100)

            avg_win = np.mean(wins) if len(wins) > 0 else 0
            avg_loss = np.mean(losses) if len(losses) > 0 else 0

            # Calculate max loss during n-day window
            max_losses_nd = []
            for buy_date in buy_signals:
                try:
                    idx_pos = tickers_returns.index.get_loc(buy_date)
                except KeyError:
                    continue

                future_nd = tickers_returns.iloc[idx_pos + 1:idx_pos + nd+1][asset]
                if len(future_nd) > 0:
                    cumulative = future_nd.cumsum() * 100
                    max_loss = cumulative.min()
                    max_losses_nd.append(max_loss)

            avg_max_loss = np.mean(max_losses_nd) if max_losses_nd else 0

            # Risk/Reward Ratio
            if avg_loss != 0:
                risk_reward_ratio = avg_win / abs(avg_loss)
            else:
                risk_reward_ratio = 0

            # Kelly Percentage
            win_rate_decimal = win_rate / 100
            if risk_reward_ratio > 0:
                kelly = win_rate_decimal - (1 - win_rate_decimal) / risk_reward_ratio
            else:
                kelly = 0

            kelly_safe = max(0, kelly)

            # Average return across all signals
            avg_return = np.mean(returns_array)

            # Strategy Return
            avg_pos = np.mean(position_sizes) if position_sizes else 0
            num_signals = len(buy_signals)

            pct_to_invest=kelly_safe

            strategy_return = avg_return * avg_pos * pct_to_invest * num_signals

            # Annualize
            annualized = strategy_return/len(tickers_returns) * 252

            asset_performance[threshold_name] = {
                'ddn_level': ddn_level,
                'count': num_signals,
                'avg_pos_size': avg_pos,
                'avg_return_nd': avg_return,
                'avg_win_nd': avg_win if len(wins) > 0 else 0,
                'avg_loss_nd': avg_loss if len(losses) > 0 else 0,
                'avg_max_loss_nd': avg_max_loss,
                'win_rate_nd': win_rate,
                'risk_reward_ratio': risk_reward_ratio,
                'kelly_percentage': kelly_safe * 100,
                'strategy_return': strategy_return,
                'annualized_return': annualized
            }

        all_results[asset] = asset_performance

    return all_results


def create_excel_report(results, output_path='optimal_buy_strategy_ndays.xlsx'):
    """
    Create Excel workbook with results
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Color scheme
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    alt_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Create sheet for each asset
    for asset, performance in results.items():
        ws = wb.create_sheet(asset)

        # Headers
        headers = ['DDN Level', 'DDN Value', 'Buy Signals', 'Avg Pos Size',
                   'Avg Ret nd (%)', 'Avg Win nd (%)', 'Avg Loss nd (%)','Avg Max Loss nd (%)',
                   'Win Rate (%)', 'Risk/Reward', 'Kelly (%)',
                   'Strategy Return', 'Annualized Return']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 16
        ws.column_dimensions['L'].width = 18
        ws.column_dimensions['M'].width = 18

        # Data rows
        row = 2
        for threshold, metrics in performance.items():
            ws.cell(row=row, column=1).value = threshold
            ws.cell(row=row, column=2).value = metrics['ddn_level']
            ws.cell(row=row, column=3).value = int(metrics['count']) if metrics['count'] > 0 else 0
            ws.cell(row=row, column=4).value = metrics['avg_pos_size']
            ws.cell(row=row, column=5).value = metrics['avg_return_nd']
            ws.cell(row=row, column=6).value = metrics['avg_win_nd']
            ws.cell(row=row, column=7).value = metrics['avg_loss_nd']
            ws.cell(row=row, column=8).value = metrics['avg_max_loss_nd']
            ws.cell(row=row, column=9).value = metrics['win_rate_nd']
            ws.cell(row=row, column=10).value = metrics['risk_reward_ratio']
            ws.cell(row=row, column=11).value = metrics['kelly_percentage']
            ws.cell(row=row, column=12).value = metrics['strategy_return']
            ws.cell(row=row, column=14).value = metrics['annualized_return']

            # Apply formatting
            for col in range(1, 14):
                cell = ws.cell(row=row, column=col)
                cell.border = border

                if col in [2, 4, 5, 6, 7, 8, 9, 10, 11, 12,13]:
                    cell.alignment = Alignment(horizontal='right')
                    if col in [2, 4]:
                        cell.number_format = '0.0000'
                    elif col in [5, 6, 7, 8, 9, 10]:
                        cell.number_format = '0.00'
                    elif col in [11, 12,13]:
                        cell.number_format = '0.00'

                if row % 2 == 0:
                    cell.fill = alt_fill

            row += 1

    # Create summary sheet
    summary_ws = wb.create_sheet('Summary', 0)
    summary_ws.column_dimensions['A'].width = 18
    summary_ws.column_dimensions['B'].width = 16
    summary_ws.column_dimensions['C'].width = 16
    summary_ws.column_dimensions['D'].width = 18

    row = 1
    summary_ws.cell(row=row, column=1).value = 'Asset'
    summary_ws.cell(row=row, column=2).value = 'Best Percentile'
    summary_ws.cell(row=row, column=3).value = 'Best Annualized Return'
    summary_ws.cell(row=row, column=4).value = 'DDN Threshold'

    for col in range(1, 5):
        cell = summary_ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    row = 2
    for asset, performance in results.items():
        best_percentile = None
        best_return = -float('inf')
        best_ddn = None

        for threshold, metrics in performance.items():
            if not np.isnan(metrics['annualized_return']):
                if metrics['annualized_return'] > best_return:
                    best_return = metrics['annualized_return']
                    best_percentile = threshold
                    best_ddn = metrics['ddn_level']

        summary_ws.cell(row=row, column=1).value = asset
        summary_ws.cell(row=row, column=2).value = best_percentile if best_percentile else 'N/A'
        summary_ws.cell(row=row, column=3).value = best_return if best_return > -float('inf') else np.nan
        summary_ws.cell(row=row, column=4).value = best_ddn if best_ddn else np.nan

        for col in range(1, 5):
            cell = summary_ws.cell(row=row, column=col)
            cell.border = border
            if col in [3, 4]:
                cell.number_format = '0.00'

        row += 1

    wb.save(output_path)
    return output_path

def create_df_report(results):
    """
    Create df's with results by tickers
    """
    ddn_levels = pd.DataFrame()
    kelly_pcts = pd.DataFrame()
    #thresholds = pd.DataFrame()

    # Create sheet for each asset
    for asset, performance in results.items():

        kelly_pcts_array=[]
        ddn_levels_array = []

        for threshold, metrics in performance.items():
            kelly_pcts_array.append(metrics['kelly_percentage'])
            ddn_levels_array.append(metrics['ddn_level'])

        kelly_pcts[asset] = kelly_pcts_array
        ddn_levels[asset] = ddn_levels_array


    return ddn_levels, kelly_pcts

# Usage:
# results = optimal_buy_strategy_ndays(tickers_returns, ddn_df, positions_df)
# create_excel_report(results)
# print("Excel file created: optimal_buy_strategy_ndays.xlsx")


if __name__ == '__main__':
    run(settings)
